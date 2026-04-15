#!/usr/bin/env python3
"""Build docs/submission/GPTHub_features_matrix.xlsx from FEATURE_MATRIX.md tables.

Лист «Матрица фич»: русские заголовки, категория (обязательные / доп.), пояснение
статуса на русском, подробное описание реализации из markdown.

Run from repo root:
  uv run --with openpyxl python scripts/build_features_xlsx.py

Regenerate whenever [FEATURE_MATRIX.md](FEATURE_MATRIX.md) changes.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[1]
MATRIX = ROOT / "FEATURE_MATRIX.md"
OUT = ROOT / "docs" / "submission" / "GPTHub_features_matrix.xlsx"

# Код статуса (как в FEATURE_MATRIX.md) → краткое пояснение по-русски для колонки E.
STATUS_RU: dict[str, str] = {
    "Implemented": (
        "Реализовано в коде: сценарий закрыт автотестами и/или живым прогоном "
        "контракта MWS (см. docs/MWS_CATALOG.md, docs/LIVE_SMOKE.md)."
    ),
    "Implemented (UI-managed)": (
        "Реализовано через штатные настройки Open WebUI (без отдельной логики "
        "в orchestrator). Обязательны переменные окружения / тумблеры из колонки "
        "«Реализация»; без них фича не включится."
    ),
    "Implemented (WOW-1)": (
        "Реализовано как дополнительная WOW-1 фича: не ломает единый поток чата, "
        "есть тесты и зафиксированный smoke; приоритет см. ROADMAP.md."
    ),
    "Implemented (WOW-3)": (
        "Реализовано как дополнительная WOW-3 фича: не ломает единый поток чата, "
        "есть тесты и записи в LIVE_SMOKE.md; приоритет см. ROADMAP.md."
    ),
    "Wired, pending live smoke": (
        "Код подключён, но нужен полный end-to-end прогон через поднятый "
        "docker-stack и запись результата в docs/LIVE_SMOKE.md."
    ),
    "Partial": "Часть пользовательских сценариев покрыта, часть — нет; см. причину в матрице.",
    "Deferred": "Не начато / отложено; в матрице указана причина или план.",
}

REFERENCE_SHEET_ROWS: list[tuple[str, str]] = [
    (
        "Implemented",
        "Код в репозитории, тесты проходят, путь подтверждён контрактом MWS "
        "или unit/integration тестом.",
    ),
    (
        "Implemented (UI-managed)",
        "Работает через встроенные возможности Open WebUI; в orchestrator отдельной "
        "реализации нет.",
    ),
    (
        "Implemented (WOW-1) / (WOW-3)",
        "Реализовано как дополнительный WOW-сценарий; требования к доказательствам "
        "те же, что у Implemented.",
    ),
    (
        "Wired, pending live smoke",
        "Код есть, нужен сквозной прогон через docker-stack.",
    ),
    ("Partial", "Сценарий закрыт не полностью."),
    ("Deferred", "Не реализовано; отложено."),
]


def parse_matrix_tables(md: str) -> list[tuple[str, str, str, str]]:
    rows: list[tuple[str, str, str, str]] = []
    in_table = False
    for line in md.splitlines():
        if not line.startswith("|"):
            in_table = False
            continue
        if re.match(r"^\|\s*Row\s*\|", line):
            in_table = True
            continue
        if not in_table:
            continue
        if re.match(r"^\|\s*---", line):
            continue
        parts = [p.strip() for p in line.strip().split("|")]
        parts = [p for p in parts if p]
        if len(parts) >= 4 and parts[0].isdigit():
            rows.append((parts[0], parts[1], parts[2], parts[3]))
    return rows


def category_for_row(row_num: str) -> str:
    n = int(row_num)
    if n <= 12:
        return "Обязательные фичи (ряды 1–12)"
    return "Дополнительный функционал (ряды 13–15)"


def status_ru(status_en: str) -> str:
    return STATUS_RU.get(
        status_en,
        f"Код «{status_en}» — см. определения в FEATURE_MATRIX.md, раздел «Статусы».",
    )


def main() -> None:
    data = parse_matrix_tables(MATRIX.read_text(encoding="utf-8"))
    if len(data) != 15:
        raise SystemExit(f"expected 15 feature rows, got {len(data)} — check {MATRIX}")

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Матрица фич"
    headers = (
        "№ ряда",
        "Категория",
        "Фича / сценарий",
        "Статус (канон, EN)",
        "Пояснение статуса (RU)",
        "Реализация в проекте (подробно, RU)",
    )
    ws.append(list(headers))
    bold = Font(bold=True)
    top_wrap = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = top_wrap

    for num, feat, stat, how in data:
        ws.append(
            [
                int(num),
                category_for_row(num),
                feat,
                stat,
                status_ru(stat),
                how,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = top_wrap

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 88

    ref = wb.create_sheet("Справка", 1)
    ref.append(
        [
            "Этот файл собран из FEATURE_MATRIX.md (корень репозитория). "
            "Пересборка: uv run --with openpyxl python scripts/build_features_xlsx.py",
        ]
    )
    ref.merge_cells("A1:B1")
    ref["A1"].font = Font(bold=True, size=11)
    ref["A1"].alignment = top_wrap
    ref.append([])
    ref.append(["Код статуса (канон)", "Расшифровка по-русски"])
    for cell in ref[3]:
        cell.font = bold
        cell.alignment = top_wrap
    for en, ru in REFERENCE_SHEET_ROWS:
        ref.append([en, ru])
    ref.column_dimensions["A"].width = 36
    ref.column_dimensions["B"].width = 72
    for row in ref.iter_rows(min_row=4, max_row=ref.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = top_wrap

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
